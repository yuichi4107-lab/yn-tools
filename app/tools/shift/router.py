"""シフト作成アプリ - ルーター"""

from datetime import date, timedelta

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import require_tool_access
from app.database import get_db
from app.users.models import User

from .models import ShiftEmployee, ShiftTemplate, ShiftSchedule, ShiftRequest, ShiftAssignment

router = APIRouter(prefix="/tools/shift", tags=["shift"])
templates = Jinja2Templates(directory="app/templates")


# ---------------------------------------------------------------------------
# メインページ
# ---------------------------------------------------------------------------

@router.get("/", response_class=HTMLResponse)
async def shift_index(
    request: Request,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """シフト管理ダッシュボード"""
    employees_q = await db.execute(
        select(ShiftEmployee)
        .where(ShiftEmployee.user_id == user.id, ShiftEmployee.is_active == True)
        .order_by(ShiftEmployee.name)
    )
    employees = employees_q.scalars().all()

    templates_q = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.user_id == user.id).order_by(ShiftTemplate.name)
    )
    shift_templates = templates_q.scalars().all()

    schedules_q = await db.execute(
        select(ShiftSchedule)
        .where(ShiftSchedule.user_id == user.id)
        .order_by(ShiftSchedule.year_month.desc())
        .limit(20)
    )
    schedules = schedules_q.scalars().all()

    # SQLAlchemyオブジェクトをJinja2 tojson用に辞書化
    employees_data = [
        {"id": e.id, "name": e.name, "role": e.role, "hourly_wage": e.hourly_wage,
         "available_days": e.available_days, "available_start": e.available_start,
         "available_end": e.available_end, "max_hours_per_week": e.max_hours_per_week}
        for e in employees
    ]
    templates_data = [
        {"id": t.id, "name": t.name, "start_time": t.start_time, "end_time": t.end_time,
         "break_minutes": t.break_minutes, "color": t.color}
        for t in shift_templates
    ]
    schedules_data = [
        {"id": s.id, "name": s.name, "year_month": s.year_month,
         "status": s.status, "ai_generated": s.ai_generated}
        for s in schedules
    ]

    return templates.TemplateResponse(
        request, "tools/shift/index.html", {
            "user": user,
            "page": "shift",
            "employees": employees_data,
            "shift_templates": templates_data,
            "schedules": schedules_data,
        }
    )


# ---------------------------------------------------------------------------
# 従業員 CRUD
# ---------------------------------------------------------------------------

@router.post("/api/employees/create")
async def api_create_employee(
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
    name: str = Form(default=""),
    role: str = Form(default="part_time"),
    hourly_wage: int = Form(default=0),
    available_days: str = Form(default="0123456"),
    available_start: str = Form(default="09:00"),
    available_end: str = Form(default="22:00"),
    max_hours_per_week: int = Form(default=40),
):
    if not name.strip():
        return {"error": "従業員名を入力してください。"}
    if role not in ("full_time", "part_time", "arbeit"):
        return {"error": "不正な役職です。"}

    emp = ShiftEmployee(
        user_id=user.id,
        name=name.strip(),
        role=role,
        hourly_wage=hourly_wage or None,
        available_days=available_days,
        available_start=available_start,
        available_end=available_end,
        max_hours_per_week=max_hours_per_week,
    )
    db.add(emp)
    await db.commit()
    return {"ok": True, "id": emp.id}


@router.post("/api/employees/update/{emp_id}")
async def api_update_employee(
    emp_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
    name: str = Form(default=""),
    role: str = Form(default="part_time"),
    hourly_wage: int = Form(default=0),
    available_days: str = Form(default="0123456"),
    available_start: str = Form(default="09:00"),
    available_end: str = Form(default="22:00"),
    max_hours_per_week: int = Form(default=40),
):
    result = await db.execute(
        select(ShiftEmployee).where(ShiftEmployee.id == emp_id, ShiftEmployee.user_id == user.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        return {"error": "従業員が見つかりません。"}
    if not name.strip():
        return {"error": "従業員名を入力してください。"}

    emp.name = name.strip()
    emp.role = role
    emp.hourly_wage = hourly_wage or None
    emp.available_days = available_days
    emp.available_start = available_start
    emp.available_end = available_end
    emp.max_hours_per_week = max_hours_per_week
    await db.commit()
    return {"ok": True}


@router.post("/api/employees/delete/{emp_id}")
async def api_delete_employee(
    emp_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ShiftEmployee).where(ShiftEmployee.id == emp_id, ShiftEmployee.user_id == user.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        return {"error": "従業員が見つかりません。"}
    emp.is_active = False
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# シフトテンプレート CRUD
# ---------------------------------------------------------------------------

@router.post("/api/templates/create")
async def api_create_template(
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
    name: str = Form(default=""),
    start_time: str = Form(default="09:00"),
    end_time: str = Form(default="18:00"),
    break_minutes: int = Form(default=60),
    color: str = Form(default="#3B82F6"),
):
    if not name.strip():
        return {"error": "テンプレート名を入力してください。"}

    tmpl = ShiftTemplate(
        user_id=user.id,
        name=name.strip(),
        start_time=start_time,
        end_time=end_time,
        break_minutes=break_minutes,
        color=color,
    )
    db.add(tmpl)
    await db.commit()
    return {"ok": True, "id": tmpl.id}


@router.post("/api/templates/update/{tmpl_id}")
async def api_update_template(
    tmpl_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
    name: str = Form(default=""),
    start_time: str = Form(default="09:00"),
    end_time: str = Form(default="18:00"),
    break_minutes: int = Form(default=60),
    color: str = Form(default="#3B82F6"),
):
    result = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.id == tmpl_id, ShiftTemplate.user_id == user.id)
    )
    tmpl = result.scalar_one_or_none()
    if not tmpl:
        return {"error": "テンプレートが見つかりません。"}
    if not name.strip():
        return {"error": "テンプレート名を入力してください。"}

    tmpl.name = name.strip()
    tmpl.start_time = start_time
    tmpl.end_time = end_time
    tmpl.break_minutes = break_minutes
    tmpl.color = color
    await db.commit()
    return {"ok": True}


@router.post("/api/templates/delete/{tmpl_id}")
async def api_delete_template(
    tmpl_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.id == tmpl_id, ShiftTemplate.user_id == user.id)
    )
    tmpl = result.scalar_one_or_none()
    if not tmpl:
        return {"error": "テンプレートが見つかりません。"}
    await db.delete(tmpl)
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# シフト表 CRUD
# ---------------------------------------------------------------------------

@router.post("/api/schedules/create")
async def api_create_schedule(
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
    name: str = Form(default=""),
    year_month: str = Form(default=""),
):
    if not name.strip():
        return {"error": "シフト表名を入力してください。"}
    if not year_month or len(year_month) != 7:
        return {"error": "年月をYYYY-MM形式で入力してください。"}

    sched = ShiftSchedule(
        user_id=user.id,
        name=name.strip(),
        year_month=year_month,
    )
    db.add(sched)
    await db.commit()
    return {"ok": True, "id": sched.id}


@router.post("/api/schedules/delete/{sched_id}")
async def api_delete_schedule(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    sched = result.scalar_one_or_none()
    if not sched:
        return {"error": "シフト表が見つかりません。"}
    # 関連する割り当て・希望も削除
    await db.execute(delete(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id))
    await db.execute(delete(ShiftRequest).where(ShiftRequest.schedule_id == sched_id))
    await db.delete(sched)
    await db.commit()
    return {"ok": True}


@router.post("/api/schedules/{sched_id}/publish")
async def api_publish_schedule(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    sched = result.scalar_one_or_none()
    if not sched:
        return {"error": "シフト表が見つかりません。"}
    sched.status = "published"
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# シフト希望 CRUD
# ---------------------------------------------------------------------------

@router.get("/api/requests/{sched_id}")
async def api_get_requests(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """指定シフト表の全希望を取得"""
    # user_idチェック: スケジュールの所有者か確認
    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    if not sched_q.scalar_one_or_none():
        return {"error": "シフト表が見つかりません。"}

    result = await db.execute(
        select(ShiftRequest).where(ShiftRequest.schedule_id == sched_id)
    )
    requests = result.scalars().all()
    return {"ok": True, "requests": [
        {
            "id": r.id,
            "employee_id": r.employee_id,
            "work_date": r.work_date.isoformat(),
            "request_type": r.request_type,
            "template_id": r.template_id,
            "note": r.note,
        } for r in requests
    ]}


@router.post("/api/requests/{sched_id}/save")
async def api_save_requests(
    sched_id: int,
    request: Request,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """シフト希望を一括保存（既存を全削除→再挿入）"""
    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    if not sched_q.scalar_one_or_none():
        return {"error": "シフト表が見つかりません。"}

    body = await request.json()
    items = body.get("requests", [])

    # 既存を削除
    await db.execute(delete(ShiftRequest).where(ShiftRequest.schedule_id == sched_id))

    # 新規挿入
    for item in items:
        req = ShiftRequest(
            schedule_id=sched_id,
            employee_id=item["employee_id"],
            work_date=date.fromisoformat(item["work_date"]),
            request_type=item["request_type"],
            template_id=item.get("template_id"),
            note=item.get("note"),
        )
        db.add(req)

    await db.commit()
    return {"ok": True, "count": len(items)}


# ---------------------------------------------------------------------------
# シフト割り当て CRUD
# ---------------------------------------------------------------------------

@router.get("/api/assignments/{sched_id}")
async def api_get_assignments(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """指定シフト表の全割り当てを取得"""
    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    if not sched_q.scalar_one_or_none():
        return {"error": "シフト表が見つかりません。"}

    result = await db.execute(
        select(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id)
    )
    assignments = result.scalars().all()
    return {"ok": True, "assignments": [
        {
            "id": a.id,
            "employee_id": a.employee_id,
            "template_id": a.template_id,
            "work_date": a.work_date.isoformat(),
            "custom_start": a.custom_start,
            "custom_end": a.custom_end,
            "note": a.note,
        } for a in assignments
    ]}


@router.post("/api/assignments/{sched_id}/save")
async def api_save_assignments(
    sched_id: int,
    request: Request,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """シフト割り当てを一括保存（既存を全削除→再挿入）"""
    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    if not sched_q.scalar_one_or_none():
        return {"error": "シフト表が見つかりません。"}

    body = await request.json()
    items = body.get("assignments", [])

    # 既存を削除
    await db.execute(delete(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id))

    # 新規挿入
    for item in items:
        assignment = ShiftAssignment(
            schedule_id=sched_id,
            employee_id=item["employee_id"],
            template_id=item.get("template_id"),
            work_date=date.fromisoformat(item["work_date"]),
            custom_start=item.get("custom_start"),
            custom_end=item.get("custom_end"),
            note=item.get("note"),
        )
        db.add(assignment)

    await db.commit()
    return {"ok": True, "count": len(items)}


# ---------------------------------------------------------------------------
# バリデーション（労基法チェック）
# ---------------------------------------------------------------------------

@router.get("/api/assignments/{sched_id}/validate")
async def api_validate_assignments(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """シフト割り当ての労基法バリデーション"""
    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    sched = sched_q.scalar_one_or_none()
    if not sched:
        return {"error": "シフト表が見つかりません。"}

    # 従業員一覧
    emp_q = await db.execute(
        select(ShiftEmployee).where(ShiftEmployee.user_id == user.id, ShiftEmployee.is_active == True)
    )
    employees = {e.id: e for e in emp_q.scalars().all()}

    # テンプレート一覧
    tmpl_q = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.user_id == user.id)
    )
    tmpl_map = {t.id: t for t in tmpl_q.scalars().all()}

    # 割り当て一覧
    assign_q = await db.execute(
        select(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id)
    )
    assignments = assign_q.scalars().all()

    warnings = []

    # 従業員ごとにチェック
    emp_assignments: dict[int, list] = {}
    for a in assignments:
        emp_assignments.setdefault(a.employee_id, []).append(a)

    for emp_id, assigns in emp_assignments.items():
        emp = employees.get(emp_id)
        if not emp:
            continue

        # 日付ソート
        assigns.sort(key=lambda x: x.work_date)

        # 週ごとの労働時間チェック
        weekly_hours: dict[str, float] = {}
        for a in assigns:
            if not a.template_id:
                continue
            tmpl = tmpl_map.get(a.template_id)
            if not tmpl:
                continue
            start_h, start_m = map(int, (a.custom_start or tmpl.start_time).split(":"))
            end_h, end_m = map(int, (a.custom_end or tmpl.end_time).split(":"))
            hours = (end_h * 60 + end_m - start_h * 60 - start_m - tmpl.break_minutes) / 60
            # ISO週番号でグループ化
            week_key = a.work_date.isocalendar()[:2]
            week_str = f"{week_key[0]}-W{week_key[1]:02d}"
            weekly_hours[week_str] = weekly_hours.get(week_str, 0) + hours

        for week_str, hours in weekly_hours.items():
            if hours > emp.max_hours_per_week:
                warnings.append(
                    f"{emp.name}: {week_str} の週間労働時間が{hours:.1f}時間（上限{emp.max_hours_per_week}時間）を超過"
                )

        # 連続勤務チェック（7日以上）
        work_dates = sorted({a.work_date for a in assigns if a.template_id})
        consecutive = 1
        for i in range(1, len(work_dates)):
            if work_dates[i] - work_dates[i - 1] == timedelta(days=1):
                consecutive += 1
                if consecutive >= 7:
                    warnings.append(
                        f"{emp.name}: {work_dates[i - 6].isoformat()}〜{work_dates[i].isoformat()} で7日以上連続勤務"
                    )
            else:
                consecutive = 1

    return {"ok": True, "warnings": warnings}


# ---------------------------------------------------------------------------
# AIシフト自動生成
# ---------------------------------------------------------------------------

@router.post("/api/schedules/{sched_id}/generate")
async def api_generate_shift(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """AIでシフトを自動生成"""
    import json
    import calendar
    from openai import AsyncOpenAI
    from app.config import settings

    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    sched = sched_q.scalar_one_or_none()
    if not sched:
        return {"error": "シフト表が見つかりません。"}

    if not settings.openai_api_key:
        return {"error": "OpenAI APIキーが設定されていません。"}

    # データ取得
    emp_q = await db.execute(
        select(ShiftEmployee).where(ShiftEmployee.user_id == user.id, ShiftEmployee.is_active == True)
    )
    employees = emp_q.scalars().all()
    if not employees:
        return {"error": "従業員が登録されていません。先に従業員を追加してください。"}

    tmpl_q = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.user_id == user.id)
    )
    shift_tmpls = tmpl_q.scalars().all()
    if not shift_tmpls:
        return {"error": "シフトテンプレートが登録されていません。先にテンプレートを追加してください。"}

    # 希望データ
    req_q = await db.execute(
        select(ShiftRequest).where(ShiftRequest.schedule_id == sched_id)
    )
    requests_list = req_q.scalars().all()

    # 年月パース
    year, month = map(int, sched.year_month.split("-"))
    num_days = calendar.monthrange(year, month)[1]
    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]

    # プロンプト構築
    emp_info = []
    for e in employees:
        days_str = "".join(weekday_names[int(d)] for d in e.available_days if d.isdigit() and int(d) < 7)
        emp_info.append(
            f"- ID:{e.id} {e.name}（{e.role}）勤務可能: {days_str} {e.available_start}〜{e.available_end}、週最大{e.max_hours_per_week}時間"
        )

    tmpl_info = []
    for t in shift_tmpls:
        h = (int(t.end_time.split(":")[0]) * 60 + int(t.end_time.split(":")[1])
             - int(t.start_time.split(":")[0]) * 60 - int(t.start_time.split(":")[1])
             - t.break_minutes) / 60
        tmpl_info.append(f"- ID:{t.id} {t.name}（{t.start_time}〜{t.end_time}、休憩{t.break_minutes}分、実働{h:.1f}時間）")

    req_info = []
    for r in requests_list:
        emp_name = next((e.name for e in employees if e.id == r.employee_id), f"ID:{r.employee_id}")
        type_label = {"day_off": "休み希望", "prefer_work": "出勤希望", "prefer_template": "シフト希望"}.get(r.request_type, r.request_type)
        req_info.append(f"- {emp_name}: {r.work_date.isoformat()} {type_label}" + (f"（備考: {r.note}）" if r.note else ""))

    dates_info = []
    for d in range(1, num_days + 1):
        dt = date(year, month, d)
        dates_info.append(f"{dt.isoformat()}({weekday_names[dt.weekday()]})")

    prompt = f"""あなたはシフト管理の専門家です。以下の情報に基づいて、{year}年{month}月のシフト表を作成してください。

## 対象日
{', '.join(dates_info)}

## 従業員
{chr(10).join(emp_info)}

## シフトテンプレート
{chr(10).join(tmpl_info)}

## シフト希望
{chr(10).join(req_info) if req_info else 'なし'}

## ルール
1. 各従業員の勤務可能曜日・時間帯を守ること
2. 週あたりの労働時間が各従業員のmax_hours_per_weekを超えないこと
3. 7日以上連続勤務を避けること（最大6連勤）
4. 休み希望（day_off）は必ず尊重すること
5. 出勤希望（prefer_work）はできるだけ尊重すること
6. シフト希望（prefer_template）がある場合は、できるだけ希望のテンプレートを割り当てること
7. 各日に適度な人数が配置されるよう、バランスよく分散すること

## 出力形式
以下のJSON形式で出力してください。他のテキストは含めないでください。
{{
  "assignments": [
    {{"employee_id": 1, "template_id": 2, "work_date": "YYYY-MM-DD"}},
    ...
  ]
}}
- 休みの日は配列に含めないでください（template_idがnullの行は不要）
- employee_idとtemplate_idは上記の従業員・テンプレートのIDを使用してください"""

    # OpenAI API呼び出し
    client = AsyncOpenAI(api_key=settings.openai_api_key)
    try:
        response = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.3,
        )
        result_text = response.choices[0].message.content
        result_data = json.loads(result_text)
    except Exception as e:
        return {"error": f"AI生成に失敗しました: {str(e)}"}

    # 結果をDBに保存（既存割り当てを置換）
    ai_assignments = result_data.get("assignments", [])
    valid_emp_ids = {e.id for e in employees}
    valid_tmpl_ids = {t.id for t in shift_tmpls}

    await db.execute(delete(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id))

    saved_count = 0
    for item in ai_assignments:
        emp_id = item.get("employee_id")
        tmpl_id = item.get("template_id")
        work_date_str = item.get("work_date")
        if not emp_id or not tmpl_id or not work_date_str:
            continue
        if emp_id not in valid_emp_ids or tmpl_id not in valid_tmpl_ids:
            continue
        try:
            wd = date.fromisoformat(work_date_str)
        except ValueError:
            continue
        assignment = ShiftAssignment(
            schedule_id=sched_id,
            employee_id=emp_id,
            template_id=tmpl_id,
            work_date=wd,
        )
        db.add(assignment)
        saved_count += 1

    sched.ai_generated = True
    await db.commit()

    # バリデーション実行
    # （保存後に自動チェック）
    assign_q = await db.execute(
        select(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id)
    )
    all_assigns = assign_q.scalars().all()
    warnings = []
    emp_map = {e.id: e for e in employees}
    tmpl_map = {t.id: t for t in shift_tmpls}

    emp_assignments: dict[int, list] = {}
    for a in all_assigns:
        emp_assignments.setdefault(a.employee_id, []).append(a)

    for emp_id, assigns in emp_assignments.items():
        emp = emp_map.get(emp_id)
        if not emp:
            continue
        assigns.sort(key=lambda x: x.work_date)

        weekly_hours: dict[str, float] = {}
        for a in assigns:
            if not a.template_id:
                continue
            tmpl = tmpl_map.get(a.template_id)
            if not tmpl:
                continue
            start_h, start_m = map(int, tmpl.start_time.split(":"))
            end_h, end_m = map(int, tmpl.end_time.split(":"))
            hours = (end_h * 60 + end_m - start_h * 60 - start_m - tmpl.break_minutes) / 60
            week_key = a.work_date.isocalendar()[:2]
            week_str = f"{week_key[0]}-W{week_key[1]:02d}"
            weekly_hours[week_str] = weekly_hours.get(week_str, 0) + hours

        for week_str, hours in weekly_hours.items():
            if hours > emp.max_hours_per_week:
                warnings.append(f"{emp.name}: {week_str} の週間労働時間が{hours:.1f}時間超過")

        work_dates = sorted({a.work_date for a in assigns if a.template_id})
        consecutive = 1
        for i in range(1, len(work_dates)):
            if work_dates[i] - work_dates[i - 1] == timedelta(days=1):
                consecutive += 1
                if consecutive >= 7:
                    warnings.append(f"{emp.name}: 7日以上連続勤務あり")
            else:
                consecutive = 1

    return {"ok": True, "count": saved_count, "warnings": warnings}


# ---------------------------------------------------------------------------
# Excelエクスポート
# ---------------------------------------------------------------------------

@router.get("/api/schedules/{sched_id}/export")
async def api_export_excel(
    sched_id: int,
    user: User = Depends(require_tool_access("shift")),
    db: AsyncSession = Depends(get_db),
):
    """シフト表をExcel(.xlsx)でエクスポート"""
    import io
    import calendar
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    sched_q = await db.execute(
        select(ShiftSchedule).where(ShiftSchedule.id == sched_id, ShiftSchedule.user_id == user.id)
    )
    sched = sched_q.scalar_one_or_none()
    if not sched:
        return {"error": "シフト表が見つかりません。"}

    # データ取得
    emp_q = await db.execute(
        select(ShiftEmployee).where(ShiftEmployee.user_id == user.id, ShiftEmployee.is_active == True).order_by(ShiftEmployee.name)
    )
    employees = emp_q.scalars().all()

    tmpl_q = await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.user_id == user.id)
    )
    tmpl_map = {t.id: t for t in tmpl_q.scalars().all()}

    assign_q = await db.execute(
        select(ShiftAssignment).where(ShiftAssignment.schedule_id == sched_id)
    )
    assignments = assign_q.scalars().all()

    # 割り当てをキー化
    assign_map: dict[tuple[int, str], ShiftAssignment] = {}
    for a in assignments:
        assign_map[(a.employee_id, a.work_date.isoformat())] = a

    # 年月パース
    year, month = map(int, sched.year_month.split("-"))
    num_days = calendar.monthrange(year, month)[1]
    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]

    # Excel生成
    wb = Workbook()
    ws = wb.active
    ws.title = sched.name[:31]

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    sat_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    sun_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 1)
    title_cell = ws.cell(row=1, column=1, value=f"{sched.name}（{year}年{month}月）")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # ヘッダー行（日付）
    ws.cell(row=2, column=1, value="従業員").font = Font(bold=True, size=10)
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).border = thin_border
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.column_dimensions[ws.cell(row=2, column=1).column_letter].width = 14

    # ヘッダー行（曜日）
    ws.cell(row=3, column=1, value="").border = thin_border

    for d in range(1, num_days + 1):
        col = d + 1
        dt = date(year, month, d)
        wd = dt.weekday()
        # 日付行
        cell = ws.cell(row=2, column=col, value=d)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        # 曜日行
        wd_cell = ws.cell(row=3, column=col, value=weekday_names[wd])
        wd_cell.border = thin_border
        wd_cell.alignment = Alignment(horizontal="center")
        wd_cell.font = Font(size=9, bold=True, color="DC2626" if wd == 6 else "2563EB" if wd == 5 else "000000")
        if wd == 5:
            wd_cell.fill = sat_fill
        elif wd == 6:
            wd_cell.fill = sun_fill
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 6

    # データ行
    for row_idx, emp in enumerate(employees, start=4):
        ws.cell(row=row_idx, column=1, value=emp.name).border = thin_border
        ws.cell(row=row_idx, column=1).font = Font(size=10)

        for d in range(1, num_days + 1):
            col = d + 1
            dt = date(year, month, d)
            key = (emp.id, dt.isoformat())
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            a = assign_map.get(key)
            if a and a.template_id:
                tmpl = tmpl_map.get(a.template_id)
                if tmpl:
                    cell.value = tmpl.name
                    # テンプレート色を背景に
                    hex_color = tmpl.color.lstrip("#")
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    # 明るい色の場合は黒文字、暗い色の場合は白文字
                    r, g, b = int(hex_color[:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
                    cell.font = Font(size=8, color="000000" if luminance > 0.5 else "FFFFFF")
            else:
                # 週末の背景色
                wd = dt.weekday()
                if wd == 5:
                    cell.fill = sat_fill
                elif wd == 6:
                    cell.fill = sun_fill

    # ストリーミングレスポンス
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"shift_{sched.year_month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
