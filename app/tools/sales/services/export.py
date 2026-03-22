"""CSV/Excel出力サービス（user_id版）"""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.tools.sales.models import SalesCompany, SalesCrmStatus

HEADERS = [
    "ID", "企業名", "住所", "電話番号", "Webサイト",
    "業種", "地域", "レビュー数", "メールアドレス",
    "お問い合わせフォーム", "SNS (Instagram)", "SNS (Twitter)", "SNS (Facebook)",
    "営業ステータス", "スコア", "メモ", "ブラックリスト",
]


async def _get_export_data(
    db: AsyncSession, user_id: int,
    status_filter: str = "", exclude_blacklist: bool = True,
) -> list[dict]:
    query = (
        select(SalesCompany)
        .join(SalesCrmStatus)
        .options(selectinload(SalesCompany.contacts), selectinload(SalesCompany.crm_status))
        .where(SalesCompany.user_id == user_id)
    )
    if status_filter:
        query = query.where(SalesCrmStatus.status == status_filter)
    if exclude_blacklist:
        query = query.where(SalesCrmStatus.is_blacklisted == False)
    query = query.order_by(SalesCrmStatus.score.desc())

    result = await db.execute(query)
    companies = result.scalars().all()

    rows = []
    for c in companies:
        emails = [ct.email for ct in c.contacts if ct.email] if c.contacts else []
        forms = [ct.contact_form_url for ct in c.contacts if ct.contact_form_url] if c.contacts else []
        first = c.contacts[0] if c.contacts else None
        rows.append({
            "ID": c.id,
            "企業名": c.name,
            "住所": c.address or "",
            "電話番号": c.phone or "",
            "Webサイト": c.website_url or "",
            "業種": c.industry or "",
            "地域": c.region or "",
            "レビュー数": c.review_count or 0,
            "メールアドレス": ", ".join(emails),
            "お問い合わせフォーム": ", ".join(forms),
            "SNS (Instagram)": first.sns_instagram or "" if first else "",
            "SNS (Twitter)": first.sns_twitter or "" if first else "",
            "SNS (Facebook)": first.sns_facebook or "" if first else "",
            "営業ステータス": c.crm_status.status if c.crm_status else "",
            "スコア": c.crm_status.score if c.crm_status else 0,
            "メモ": c.crm_status.memo or "" if c.crm_status else "",
            "ブラックリスト": "○" if (c.crm_status and c.crm_status.is_blacklisted) else "",
        })
    return rows


async def export_csv(
    db: AsyncSession, user_id: int,
    status_filter: str = "", exclude_blacklist: bool = True,
) -> str:
    rows = await _get_export_data(db, user_id, status_filter, exclude_blacklist)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=HEADERS)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


async def export_excel(
    db: AsyncSession, user_id: int,
    status_filter: str = "", exclude_blacklist: bool = True,
) -> bytes:
    rows = await _get_export_data(db, user_id, status_filter, exclude_blacklist)

    wb = Workbook()
    ws = wb.active
    ws.title = "企業リスト"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=ri, column=ci, value=row[h])
            cell.border = border

    widths = {
        "ID": 6, "企業名": 30, "住所": 35, "電話番号": 16, "Webサイト": 30,
        "業種": 15, "地域": 15, "レビュー数": 10, "メールアドレス": 30,
        "お問い合わせフォーム": 30, "SNS (Instagram)": 25, "SNS (Twitter)": 25,
        "SNS (Facebook)": 25, "営業ステータス": 12, "スコア": 8, "メモ": 25,
        "ブラックリスト": 12,
    }
    for ci, h in enumerate(HEADERS, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(h, 15)

    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
